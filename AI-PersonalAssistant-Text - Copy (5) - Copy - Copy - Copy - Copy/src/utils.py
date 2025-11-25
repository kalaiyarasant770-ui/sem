
#✅ Intent classification, workflow nodes
# utils.py
import json
import logging
import time
import pandas as pd
import numpy as np
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any, Optional
from openai import OpenAI

from config import (
    OPENAI_API_KEY, OPENAI_MODEL, TARIFF_RATE, COLUMN_MAPPING
)
from config import COLUMN_MAPPING
from database import get_tables_for_time_range, engine, text

# Initialize OpenAI client
t_llm_init = time.time()
logging.info("🔄 Initializing OpenAI client...")
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
if openai_client:
    logging.info(f"✅ OpenAI client initialized in {time.time() - t_llm_init:.2f} seconds")
else:
    logging.warning("⚠️  OpenAI API key not found - LLM features disabled")

# ==================== METRICS CALCULATIONS ====================

def _estimate_energy_from_power(df: pd.DataFrame, device_id: int, power_col: str = 'power') -> float:
    """Calculate energy consumption from power readings using trapezoidal integration."""
    logging.info(f"\n{'='*60}")
    logging.info(f"⚡ ENERGY CALCULATION START")
    logging.info(f"Device ID: {device_id}")
    logging.info(f"{'='*60}")
    
    if power_col not in df.columns or df[power_col].isna().all():
        logging.warning(f"❌ No valid power data for device_id: {device_id}")
        logging.info(f"{'='*60}\n")
        return 0.0
    
    time_column = 'datetime' if 'datetime' in df.columns else 'updatedtime'
    df = df.sort_values(time_column).copy()
    df[power_col] = pd.to_numeric(df[power_col], errors='coerce')
    df = df.dropna(subset=[power_col])
    
    if df.empty or len(df) < 2:
        logging.warning(f"❌ Insufficient data points: {len(df)}")
        logging.info(f"{'='*60}\n")
        return 0.0
    
    logging.info(f"📊 Data points: {len(df)}")
    logging.info(f"📊 Time range: {df[time_column].iloc[0]} to {df[time_column].iloc[-1]}")
    
    # Calculate time differences in hours
    time_diffs = df[time_column].diff().dt.total_seconds() / 3600
    power_values = df[power_col].values
    
    # Trapezoidal integration
    valid_mask = ~time_diffs.isna()
    time_diffs_clean = time_diffs[valid_mask].values
    power_avg = (power_values[:-1][valid_mask[1:].values] + power_values[1:][valid_mask[1:].values]) / 2
    energy_kwh = np.sum(power_avg * time_diffs_clean) if len(power_avg) > 0 else 0.0
    #energy_kwh = np.sum(power_values[:-1] * time_diffs_clean)  # Use power[i] × Δt[i]

    logging.info(f"⚡ Estimated energy: {energy_kwh:.2f} kWh")
    logging.info(f"💰 Cost (@ ₹{TARIFF_RATE}/kWh): ₹{energy_kwh * TARIFF_RATE:.2f}")
    logging.info(f"{'='*60}\n")
    return energy_kwh

def get_efficiency_rating(cop: Optional[float]) -> str:
    """Get efficiency rating based on COP value."""
    if cop is None:
        return "N/A"
    if cop > 4.0:
        return "Excellent"
    if cop > 3.0:
        return "Good"
    return "Poor"

# ==================== EXCEL REPORT GENERATION ====================

def get_actually_used_pins(device_id: int, time_info: Dict[str, Any]) -> set:
    """
    Check first 2-3 rows of data to see which pins actually exist in the JSON.
    Returns: set of pins like {'V14', 'V15', 'V18', ...}
    """
    logging.info(f"\n{'='*60}")
    logging.info(f"🔍 DETECTING ACTUALLY USED PINS")
    logging.info(f"Device ID: {device_id}")
    logging.info(f"{'='*60}")
    
    tables = get_tables_for_time_range(time_info)
    if not tables:
        logging.warning("⚠️  No tables available")
        return set()
    
    used_pins = set()
    
    # Check first table (most recent data)
    table = tables[0]
    json_column = 'device_value' if table == "device_data" else 'raw_value'
    time_column = 'datetime' if table == "device_data" else 'updatedtime'
    
    sql = f"""
    SELECT {json_column}
    FROM {table}
    WHERE device_id = :device_id
    ORDER BY {time_column} DESC
    LIMIT 3
    """
    try:
        with engine.connect() as conn:
            result = conn.execute(text(sql), {"device_id": device_id}).fetchall()
        
        if not result:
            logging.warning(f"⚠️  No data rows found in {table}")
            return used_pins
        
        logging.info(f"Checking {len(result)} sample rows from {table}")
        
        for idx, row in enumerate(result, 1):
            json_str = row[0]
            if json_str:
                try:
                    data = json.loads(json_str)
                    # Get all keys that start with 'V'
                    pins_in_row = {k for k in data.keys() if k.startswith('V')}
                    used_pins.update(pins_in_row)
                    logging.info(f"   Row {idx}: Found {len(pins_in_row)} pins")
                except json.JSONDecodeError:
                    logging.warning(f"   Row {idx}: Failed to parse JSON")
        
        logging.info(f"✅ Total unique pins found: {len(used_pins)}")
        logging.info(f"   Pins: {sorted(used_pins, key=lambda x: int(x[1:]) if x[1:].isdigit() else 999)}")
        logging.info(f"{'='*60}\n")
        
    except Exception as e:
        logging.error(f"❌ Failed to detect pins: {e}")
    
    return used_pins

def get_device_pin_mapping(cluster_id: int, device_id: int, used_pins: set = None) -> Dict[str, str]:
    """
    Fetch pin-to-name mapping from datastream table ONLY for pins that are actually used.
    
    Args:
        cluster_id: Cluster ID
        device_id: Device ID
        used_pins: Set of pins to map (e.g., {'V14', 'V15', 'V18'}). If None, maps all pins.
    
    Returns: {'V20': 'CH Return Temp', 'V19': 'CH Leaving Temp', ...}
    """
    logging.info(f"\n{'='*60}")
    logging.info(f"📌 FETCHING PIN MAPPING")
    logging.info(f"Cluster ID: {cluster_id}, Device ID: {device_id}")
    if used_pins:
        logging.info(f"Filtering for {len(used_pins)} used pins")
    logging.info(f"{'='*60}")
    
    sql = """
    SELECT pin, name, address
    FROM datastream
    WHERE cluster_id = :cluster_id AND isdelete = 0
    ORDER BY pin
    """
    
    pin_mapping = {}
    
    try:
        with engine.connect() as conn:
            result = conn.execute(text(sql), {"cluster_id": cluster_id}).fetchall()
        
        if not result:
            logging.warning(f"⚠️  No datastream found for cluster_id={cluster_id}")
            return pin_mapping
        
        logging.info(f"Found {len(result)} datastream entries")
        
        for row in result:
            pin = row[0]  # V20, V19, etc.
            
            # ✅ SKIP if this pin is not in the used_pins set
            if used_pins is not None and pin not in used_pins:
                continue
            
            name = row[1]  # Fallback name
            address_json = row[2]  # JSON string (ARRAY format!)
            
            pin_name = None
            
            # Parse address as JSON ARRAY first
            if address_json:
                try:
                    address_array = json.loads(address_json)
                    
                    if isinstance(address_array, list) and len(address_array) > 0:
                        # ✅ NEW: Iterate through ALL objects in the array
                        for address_obj in address_array:
                            if not isinstance(address_obj, dict):
                                continue
                            
                            # Check if this object's pin matches
                            obj_pin = address_obj.get('pin')
                            if obj_pin != pin:
                                continue
                            
                            # ✅ Check address-1, address-2, address-3, etc. IN ORDER
                            address_keys = sorted([k for k in address_obj.keys() if k.startswith('address-')])
                            
                            for addr_key in address_keys:
                                addr_data = address_obj[addr_key]
                                
                                if isinstance(addr_data, dict) and 'devices' in addr_data:
                                    devices_list = addr_data.get('devices', [])
                                    
                                    # ✅ Check if our device_id is in this address's devices
                                    if isinstance(devices_list, list) and device_id in devices_list:
                                        pin_name = addr_data.get('params')
                                        if pin_name:
                                            logging.info(f"   ✅ {pin} → {pin_name} (from {addr_key})")
                                            break  # Found in this address, stop checking others
                            
                            # If found pin_name in any address-X, stop checking other objects
                            if pin_name:
                                break
                
                except json.JSONDecodeError:
                    logging.warning(f"   ⚠️  Failed to parse JSON for pin {pin}")
            
            # Fallback to 'name' column if no address match
            if not pin_name and name:
                pin_name = name
                logging.info(f"   📝 {pin} → {pin_name} (from name column)")
            
            # Add to mapping
            if pin_name:
                pin_mapping[pin] = pin_name
            else:
                logging.warning(f"   ⚠️  No name found for pin {pin}, using pin as-is")
                pin_mapping[pin] = pin
        
        logging.info(f"✅ Total pins mapped: {len(pin_mapping)}")
        logging.info(f"{'='*60}\n")
        
    except Exception as e:
        logging.error(f"❌ Failed to fetch pin mapping: {e}")
    
    return pin_mapping


def generate_excel_report_in_memory(device_id: int, time_info: dict, user_query: str, device_name: str, device_type: int = 1, cluster_id: int = None) -> dict:
    """Generate Excel report in memory without saving to disk."""
    logging.info(f"\n{'='*80}")
    logging.info(f"📊 EXCEL REPORT GENERATION START")
    logging.info(f"{'='*80}")
    logging.info(f"Device ID: {device_id}")
    logging.info(f"Device Type: {device_type}")
    logging.info(f"Cluster ID: {cluster_id}")
    logging.info(f"Time range: {time_info.get('start_time')} to {time_info.get('end_time')}")
    
    t_start = time.time()
    
    start_dt = datetime.strptime(time_info['start_time'], '%Y-%m-%dT%H:%M:%S')
    end_dt = datetime.strptime(time_info['end_time'], '%Y-%m-%dT%H:%M:%S')
    
    tables = get_tables_for_time_range(time_info)
    if not tables:
        logging.error("❌ No tables found for time range")
        logging.info(f"{'='*80}\n")
        return {'excel_bytes': None, 'row_count': 0, 'success': False}
    
    logging.info(f"📋 Querying {len(tables)} tables")
    
    all_data = []
    
    for idx, table in enumerate(tables, 1):
        logging.info(f"[{idx}/{len(tables)}] Querying table: {table}")
        json_column = 'device_value' if table == "device_data" else 'raw_value'
        time_column = 'datetime' if table == "device_data" else 'updatedtime'
        
        sql = f"""
        SELECT device_id, {time_column}, {json_column}
        FROM {table}
        WHERE device_id = :device_id 
        AND {time_column} BETWEEN :start AND :end
        ORDER BY {time_column} ASC
        """
        
        params = {
            'device_id': device_id,
            'start': time_info['start_time'],
            'end': time_info['end_time']
        }
        try:
            with engine.connect() as connection:
                result = connection.execute(text(sql), params).fetchall()
                if result:
                    df = pd.DataFrame(result)
                    df.columns = ['device_id', 'timestamp', 'json_data']
                    all_data.append(df)
                    logging.info(f"   ✅ Fetched {len(df)} rows from {table}")
                else:
                    logging.info(f"   ⚠️  No data in {table}")
        except Exception as e:
            if "doesn't exist" not in str(e):
                logging.error(f"   ❌ Error querying {table}: {e}")
            else:
                logging.warning(f"   ⚠️  Table {table} doesn't exist (skipping)")
            continue
    
    if not all_data:
        logging.error("❌ No data found in any table")
        logging.info(f"{'='*80}\n")
        return {'excel_bytes': None, 'row_count': 0, 'success': False}
    
    df_combined = pd.concat(all_data, ignore_index=True)
    df_combined['timestamp'] = pd.to_datetime(df_combined['timestamp'])
    df_combined = df_combined.sort_values('timestamp')
    
    logging.info(f"📊 Total rows fetched: {len(df_combined)}")
    
    def parse_json_safe(json_str):
        try:
            return json.loads(json_str) if json_str else {}
        except:
            return {}
    
    df_combined['parsed_data'] = df_combined['json_data'].apply(parse_json_safe)

    # ✅ STEP 1: Detect actually used pins from data
    used_pins = get_actually_used_pins(device_id, time_info)
    
    # ✅ STEP 2: Get pin mapping ONLY for used pins
    if cluster_id and used_pins:
        column_mapping = get_device_pin_mapping(cluster_id, device_id, used_pins=used_pins)
        logging.info(f"📌 Using filtered pin mapping: {len(column_mapping)} pins (from {len(used_pins)} detected)")
    elif cluster_id:
        column_mapping = get_device_pin_mapping(cluster_id, device_id, used_pins=None)
        logging.info(f"📌 Using full pin mapping: {len(column_mapping)} pins")
    else:
        logging.warning("⚠️  No cluster_id provided, using raw pin names")
        column_mapping = {}
    
    # Group by hour
    logging.info(f"🕐 Grouping data by hour...")
    hourly_data = []
    current_hour = start_dt.replace(minute=0, second=0, microsecond=0)
    
    hour_count = 0
    while current_hour <= end_dt:
        next_hour = current_hour + timedelta(hours=1)
        
        valid_row = None
        valid_timestamp = None
        found_exact = False
        
        # ✅ STEP 1: Try to find data at EXACT hour mark (e.g., 16:00:00)
        exact_time_df = df_combined[df_combined['timestamp'] == current_hour]
        
        if not exact_time_df.empty:
            for idx in exact_time_df.index:
                row_data = exact_time_df.loc[idx, 'parsed_data']
                if row_data and any(v != '0' and v != 0 and v != '' for v in row_data.values() if v):
                    valid_row = row_data
                    valid_timestamp = current_hour  # Use exact hour
                    found_exact = True
                    logging.debug(f"   ✅ Exact: {current_hour}")
                    break
        
        # ✅ STEP 2: If no exact match, find MOST RECENT row in PREVIOUS hour window
        if not valid_row:
            # Look for data between previous_hour and current_hour (excluding current_hour)
            previous_hour = current_hour - timedelta(hours=1)
            prev_hour_df = df_combined[
                (df_combined['timestamp'] > previous_hour) & 
                (df_combined['timestamp'] < current_hour)
            ]
            
            if not prev_hour_df.empty:
                # Sort by timestamp descending to get the most recent
                prev_hour_df_sorted = prev_hour_df.sort_values('timestamp', ascending=False)
                
                for idx in prev_hour_df_sorted.index:
                    row_data = prev_hour_df_sorted.loc[idx, 'parsed_data']
                    if row_data and any(v != '0' and v != 0 and v != '' for v in row_data.values() if v):
                        valid_row = row_data
                        actual_time = prev_hour_df_sorted.loc[idx, 'timestamp']
                        valid_timestamp = current_hour  # ✅ Display as exact hour
                        logging.debug(f"   ⏰ Fallback: {actual_time} → displayed as {current_hour}")
                        break
        
        # ✅ STEP 3: Add to report if we found data
        if valid_row and valid_timestamp:
            hour_count += 1
            row_dict = {
                'Device Name': device_name,
                'Time': valid_timestamp.strftime('%b %d, %Y, %H:%M:%S')  # Always shows exact hour
            }
            
            if column_mapping:
                # Use dynamic pin mapping (ONLY for pins that exist in data)
                for pin, pin_name in column_mapping.items():
                    value = valid_row.get(pin, 0)
                    
                    # ✅ Handle nested dictionaries (convert to string for Excel)
                    if isinstance(value, dict):
                        value = ', '.join(f"{k}:{v}" for k, v in value.items())
                    # Convert to proper type
                    elif isinstance(value, str):
                        try:
                            value = float(value) if '.' in value else int(value)
                        except:
                            pass
                    row_dict[pin_name] = value
            else:
                # If no mapping, add all JSON keys as-is (filtered by used_pins)
                for key, value in valid_row.items():
                    if key != 'slave_id':  # Skip slave_id
                        # Only include if in used_pins (if available)
                        if not used_pins or key in used_pins:
                            # ✅ Handle nested dictionaries
                            if isinstance(value, dict):
                                value = ', '.join(f"{k}:{v}" for k, v in value.items())
                            row_dict[key] = value
            
            hourly_data.append(row_dict)
        
        current_hour = next_hour
    
    logging.info(f"   Processed {hour_count} hours with valid data")
    
    if not hourly_data:
        logging.error("❌ No valid hourly data found")
        logging.info(f"{'='*80}\n")
        return {'excel_bytes': None, 'row_count': 0, 'success': False}
    
    df_report = pd.DataFrame(hourly_data)
    logging.info(f"📊 Generated report with {len(df_report)} rows and {len(df_report.columns)} columns")
    
    # Create Excel in memory
    logging.info(f"📝 Creating Excel workbook...")
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{device_name} Report"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = list(df_report.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(df_report.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    
    excel_bytes = output.getvalue()
    file_size_kb = len(excel_bytes) / 1024
    
    logging.info(f"✅ Excel generated successfully:")
    logging.info(f"   Size: {file_size_kb:.2f} KB")
    logging.info(f"   Rows: {len(df_report)}")
    logging.info(f"   Columns: {len(df_report.columns)}")
    logging.info(f"   Time taken: {time.time() - t_start:.2f} seconds")
    logging.info(f"{'='*80}\n")
    
    return {
        'excel_bytes': excel_bytes,
        'row_count': len(df_report),
        'success': True
    }

# ==================== LLM HELPER FUNCTIONS ====================

def parse_recommendation_response(user_message: str, last_recommendation: str) -> Dict[str, Any]:
    """Use LLM to determine if user is accepting/rejecting recommendation."""
    if not last_recommendation or not openai_client:
        return {'action': 'new_query', 'query': user_message}
    
    logging.debug(f"🤖 Parsing user response to recommendation...")
    
    prompt = f"""
    Determine user's intent.
    
    Previous Recommendation: "{last_recommendation}"
    User's Response: "{user_message}"
    
    1. ACCEPT - User wants to proceed ("yes", "ok", "sure", "go ahead", "give me")
    2. REJECT - User doesn't want it ("no", "nope", "skip")
    3. NEW_QUERY - User asking something different
    
    Return JSON:
    {{
        "action": "accept" or "reject" or "new_query",
        "extracted_query": "full recommendation if accepted",
        "reasoning": "brief explanation"
    }}
    """
    
    try:
        response = openai_client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            max_tokens=150,
            temperature=0
        )
        
        result = json.loads(response.choices[0].message.content)
        action = result.get('action', 'new_query')
        reasoning = result.get('reasoning', '')
        
        logging.info(f"🤖 LLM determined action: {action} - {reasoning}")
        
        if action == 'accept':
            extracted_query = result.get('extracted_query', last_recommendation)
            return {'action': 'accept', 'query': extracted_query.replace('?', '').strip()}
        elif action == 'reject':
            return {'action': 'reject', 'query': None}
        else:
            return {'action': 'new_query', 'query': user_message}
    except Exception as e:
        logging.error(f"❌ LLM parsing failed: {e}, treating as new query")
        return {'action': 'new_query', 'query': user_message}


def _get_efficiency_note(rating: str) -> str:
    """Generate note based on efficiency rating."""
    if rating == "Excellent":
        return "Operating at optimal efficiency."
    elif rating == "Good":
        return "Performance is within acceptable range."
    elif rating == "Poor":
        return "Need maintenance for chiller."
    else:
        return "Efficiency data unavailable."